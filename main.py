import pandas as pd
import glob
import openpyxl
from fpdf import FPDF
from pathlib import Path

filepaths=glob.glob("Invoices/*.xlsx")

for filepath in filepaths:

    pdf=FPDF(orientation="P", unit="mm", format="A4")

    filename=Path(filepath).stem
    copy=filename
    in_nr, date=copy.split("-")

    pdf.add_page()
    pdf.set_font(family="Arial", style="B", size=16)
    pdf.cell(w=20, h=10, txt=f"Invoice nr. {in_nr}", ln=1)
    pdf.cell(w=20, h=10, txt=f"Date: {date}", ln=1)

    df = pd.read_excel(filepath, sheet_name="Sheet 1")
    columns=df.columns
    columns=[item.replace("_", " ").title() for item in columns]

    pdf.set_font(family="Times", size=12, style="B")
    pdf.cell(w=28, h=10, txt=columns[0], border=1)
    pdf.cell(w=55, h=10, txt=columns[1], border=1)
    pdf.cell(w=38, h=10, txt=columns[2], border=1)
    pdf.cell(w=40, h=10, txt=columns[3], border=1)
    pdf.cell(w=32, h=10, txt=columns[4], ln=1, border=1)

    total=0
    for index, row in df.iterrows():
        pdf.set_font(family="Times", size=12)
        pdf.cell(w=28, h=10, txt=str(row['product_id']), border=1)
        pdf.cell(w=55, h=10, txt=str(row['product_name']), border=1)
        pdf.cell(w=38, h=10, txt=str(row['amount_purchased']), border=1)
        pdf.cell(w=40, h=10, txt=str(row['price_per_unit']), border=1)
        pdf.cell(w=32, h=10, txt=str(row['total_price']), ln=1, border=1)
        total=total+row["total_price"]

    pdf.set_font(family="Times", size=12)
    pdf.cell(w=28, h=10, txt="", border=1)
    pdf.cell(w=55, h=10, txt="", border=1)
    pdf.cell(w=38, h=10, txt="", border=1)
    pdf.cell(w=40, h=10, txt="", border=1)
    pdf.cell(w=32, h=10, txt=str(total), ln=1, border=1)

    pdf.set_font(family="Times", size=12, style="B")
    pdf.cell(w=0, h=10, txt=f"The total price is {total}", ln=1)
    pdf.set_font(family="Times", size=14, style="B")
    pdf.cell(w=26, h=10, txt="PythonHow")
    pdf.image("pythonhow.png", w=10)

    pdf.output(f"PDFs/{filename}.pdf")